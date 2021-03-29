from django.shortcuts import render ,redirect
from .forms import UserLoginForm 
from django.contrib import messages
from django.http import HttpResponse ,HttpResponseNotFound 
from django.contrib.auth import  login, authenticate,logout
from . filterUnAvailableSemesters import filterUnAvailableSemesters
from . encryptdecrypt import encrypt ,decrypt 
from django.conf import settings
from .forms import UserLoginForm ,UserRegisterForm ,UploadedScoreForm
from django.contrib.auth.decorators import login_required
import json
import requests
import time  
import openpyxl
import base64
from Cryptodome.Cipher import AES

from openpyxl.styles import Protection ,Font, Color, Alignment, Border, Side, colors
from django.core.files.storage import FileSystemStorage
from . generatescorelist import generatescorelist ,validatelist 
from . basicunit import MergebasicScorelist
# Create your views here.
def landing(request) :
   return render(request,'GradeManagerapp/landing.html')


def login_view(request):
  next = request.GET.get('next')
  if request.method == 'POST':
    print(request.POST['serverprogtypeApi'])
   
    settings.BASE_URL=settings.CIPHER["BASE_URL"+request.POST['serverprogtypeApi']]

    #print(request.POST['serverprogtypeApi'].value)
    form = UserLoginForm(request.POST) 
    if form.is_valid():

       username = form.cleaned_data.get('username')
       password = form.cleaned_data.get('password')
       user = authenticate(username=username,password=password)
       if not user :
               #raise forms.ValidationError('This user does not exist')
               print('This user does not exist 4')
               messages.error(request,"This user does not exist")
               
       if not user.check_password(password):
                  #raise forms.ValidationError('In Correct password')
                  print('This user does not exist 5') 
                  messages.error(request,"In Correct password")
                  
       if not user.is_active:
                  #raise forms.ValidationError('This user is not active')
                   print('This user does not exist 6')
                   messages.error(request,"This user is not active")
                   
       login(request,user)
       if next :
            return redirect(next)

       request.session['serverprogtypeApi'] =request.POST['serverprogtypeApi']
       print(request.session['serverprogtypeApi']) 
       return redirect('landing')
    else:
        for error in form.non_field_errors() :
            messages.error(request,error )  
            form = UserLoginForm()

    context ={'form' : form }
    return render(request,'GradeManagerapp/login_view.html',context)

  context ={'form' : UserLoginForm() }
  return render(request,'GradeManagerapp/login_view.html',context)


def logout_view(request):
    return render(request,'GradeManagerapp/logout_view.html')

@login_required
def processcourses_view(request):
    print("Inside processcourses_view")
    courselist=""
    time.sleep(1)
    api=settings.BASE_URL+'/api/Camp/PythonGetAvailableCoursesForEmail'
    print(api)
    try:
         
          params={'email':request.user.email}       
          r = requests.get(api,params)
          r.raise_for_status()
          print(r.text)
          courselist = json.loads(r.text)
          if len(courselist) > 0 :          
            courselist = filterUnAvailableSemesters(courselist)   
          messages.success(request, str (len(courselist))+ " Course(s) Successfully Loaded")

    except requests.exceptions.HTTPError as errh:
         messages.error (request,"Problem Loading Courses :=> "+errh.response.text)
    except requests.exceptions.ConnectionError as errc:
         messages.error (request,"Problem Loading Courses :=> "+errc)

    except requests.exceptions.Timeout as timeout:
    # Maybe set up for a retry, or continue in a retry loop
          messages.error(request,"Problem Loading Courses :=> "+ timeout)
    
    except requests.exceptions.TooManyRedirects as manyredirect:
    # Tell the user their URL was bad and try a different one
          messages.error(request,"Problem Loading Courses :=> "+ manyredirect)
    except requests.exceptions.RequestException as requestexception:
    # catastrophic error. bail.
          messages.error(request,"Problem Loading Courses :=> "+ manyredirect)
    except  Exception as inst:
          print(inst)
          messages.error(request,"Problem Loading Courses :=> "+ inst)
               
    return render(request,'GradeManagerapp/processcourses_view.html',{'courselist':courselist,'days':range(1, 32),'months':range(1, 13),})


@login_required
def displayCourse_view(request):  
    courselist={}
    theheader={}
    if request.method=='POST':
        print("A Post Message , Details Below")
        print( request.POST)
  
       
        print("A[B]===============")


        includescoretemp = request.POST.get('includescore','False')
        includescore=False
        if includescoretemp=='on': 
            includescore=True

        orderbymatricnotemp = request.POST.get('orderbymatricno','False')
        orderbymatricno=False
        if orderbymatricnotemp=='on': 
            orderbymatricno=True
        crsid = request.POST['Selectedcourse']
        year = request.POST['year']
        month = request.POST['month']
        day = request.POST['day']
        reportname='Score_Sheet_Printing'
        step='normal'

        params = {'includescore':includescore,'longerreporttype':crsid,'orderbymatricno':orderbymatricno,'reportname':reportname,'step':step,'year':year,'month':month,'day':day}
        api=api=settings.BASE_URL+'/api/Student/PythonPullForscoreEntryUsingCrsGuid'
       
        print(api)
        try:
            headers = {'content-type': 'application/json'}
            r = requests.post(api,json=params,headers=headers)
            if (r.text=='[]'):
                  messages.error(request, "No Course was downloaded, Check Report Constraints")
                  return render (request,'GradeManagerapp/displayCourse_view.html',{'courselist':courselist})
            courselist = json.loads(r.text)
            theheader={
                    "AsessionId":courselist[0]['MYASESSIONID'],
                    "SemesterId":courselist[0]['MYSEMESTERID'],
                    "LevelToDo":courselist[0]['MYLEVELTODO'],
                    "CourseState":courselist[0]['MYCOURSESTATE'],
                    "CourseUnit":courselist[0]['MYCOURSEUNIT'],
                    "CourseNature":courselist[0]['MYCOURSENATURE'],
                    "AsetId":courselist[0]['MYASETID'],
                    "CourseId":courselist[0]['MYCOURSEID'],

            }
            request.session['courselist'] = courselist
            request.session['params'] = params
            messages.success(request, str (len(courselist))+ " Students Successfully Loaded")
            
        except  Exception as inst:
            print("See Error Details Below /n")
            print(r.text)
            print(type(r.text))
            print(courselist)
            print(inst)
            messages.error(request, "Error Check Connection or Contact Admin")
    return render (request,'GradeManagerapp/displayCourse_view.html',{'courselist':courselist})



@login_required
def downloadScoresheet_xls(request):
  courselist={}
  if request.method=='GET':
       if request.session.has_key('courselist'):
                
                courselist = request.session['courselist']
                filename= courselist[0]['MYCOURSEID']+courselist[0]['MYASESSIONID'].replace("/", "_") +courselist[0]['MYSEMESTERID']+'.xls'
                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = "attachment; filename="+filename
           
                workbook = openpyxl.Workbook()
                worksheet=workbook.active
                worksheet.protection.sheet = True
                worksheet.protection.enable()

                bold_font = Font(bold=True)
                big_red_text = Font(bold=True, color="ff0000", size=14)
                center_aligned_text = Alignment(horizontal="center")

               
                workbook.security.set_workbook_password(settings.WORKBOOKHASHED_PASSWORD, already_hashed=False)
                worksheet.protection.password = settings.CIPHER_PASS
                if request.session.has_key('params'):
                    params = request.session['params']
                else:
                    print('params was NOT passed in session')  
                worksheet.title='SCORESSHEET'
                col_start=0
                row_start=7
                
                SECRETKEY={
                    "CCODE":settings.CIPHER_PASS,
                    "TOTAL":str(len(courselist)),
                    "EMAIL":request.user.email,
                    "CID":params['longerreporttype'],
                    "AsessionId":courselist[0]['MYASESSIONID'],
                    "SemesterId":courselist[0]['MYSEMESTERID'],
                    "LevelToDo":courselist[0]['MYLEVELTODO'],
                    "CourseState":courselist[0]['MYCOURSESTATE'],
                    "CourseUnit":courselist[0]['MYCOURSEUNIT'],
                    "CourseNature":courselist[0]['MYCOURSENATURE'],
                    "AsetId":courselist[0]['MYASETID'],
                    "CourseId":courselist[0]['MYCOURSEID'],

                }

                SECRETKEY_STR = json.dumps(SECRETKEY) 
                #SECRETKEY_STR= SECRETKEY_STR.rjust(320, 'X')
                # cipher = AES.new(settings.CIPHER_PASS.rjust(32, 'X'),AES.MODE_ECB) # never use ECB in strong systems obviously
                
                # SECRETKEYciphertext = base64.b64encode(cipher.encrypt(SECRETKEY_STR))
                # print(SECRETKEYciphertext)
                # decoded = cipher.decrypt(base64.b64decode(SECRETKEYciphertext))
                                
                # # key = get_random_bytes(16)
                # # cipher = AES.new(key, AES.MODE_EAX)
                # # ciphertext, tag = cipher.encrypt_and_digest(SECRETKEY_STR)

                # # print(key)
                # # print(cipher)
                # # print(ciphertext)

                encrypted= encrypt(SECRETKEY_STR,settings.CIPHER_PASS) 
                # decrypted= decrypt(encrypted,settings.CIPHER_PASS)
                # print("All above1")
                # print(encrypted)
                # print("All above2")
                # print(decrypted)
                # print("All above3")
                
                SECRETKEYciphertext=str(encrypted)
                print(SECRETKEYciphertext)
                worksheet.cell(1 ,1).value=SECRETKEYciphertext
                worksheet.cell(1 ,1).font=Font(color="ffffff", size=2)
               
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 20
                worksheet.column_dimensions['D'].width = 20
                worksheet.column_dimensions['E'].width = 20
                now = time.strftime("%x")  
                worksheet.cell(1 ,3).value = now  
                worksheet.cell(1 ,3).font=Font(color="ffffff", size=2)
                worksheet.cell(3 ,2).value='THE POLYTECHNIC IBADAN'
                worksheet.merge_cells('B3:E3')
                worksheet["B3"].font = big_red_text
                worksheet["B3"].alignment = center_aligned_text

                worksheet.cell(4 ,2).value='INTERNAL RESULT DOCUMENT'
                worksheet.merge_cells('B4:E4')
                worksheet["B4"].font = big_red_text
                worksheet["B4"].alignment = center_aligned_text

                worksheet.cell(5 ,2).value='COURSE CODE : '+courselist[0]['MYCOURSEID']+' SESSION : '+courselist[0]['MYASESSIONID'] +' SEMESTER : '+courselist[0]['MYSEMESTERID'] 
                worksheet.merge_cells('B5:E5')

                worksheet["B5"].font = big_red_text
                worksheet["B5"].alignment = center_aligned_text
                
                for index,row in  enumerate(courselist):               
                    worksheet.cell(index+row_start ,col_start+1).value= index+1
                    worksheet.cell(index+row_start ,col_start+1).alignment = center_aligned_text
                    worksheet.cell(index+row_start ,col_start+2).font=bold_font 
                    worksheet.cell(index+row_start ,col_start+2).value= row['MYSTUDENTID']
                    worksheet.cell(index+row_start ,col_start+3).value= row['MYSURNAME']
                    worksheet.cell(index+row_start ,col_start+4).value= row['MYMIDDLENAME']
                    worksheet.cell(index+row_start ,col_start+5 ).value = row['MYFIRSTNAME']

                    worksheet.cell(index+row_start ,col_start+8).value= row['MYCOURSEID']
                    worksheet.cell(index+row_start ,col_start+8).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+9 ).value = row['MYSCORESHEETCLASSID']
                    worksheet.cell(index+row_start ,col_start+9).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+10).value= row['MYCOURSESTATE']
                    worksheet.cell(index+row_start ,col_start+10).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+11 ).value = row['MYCOURSENATURE']
                    worksheet.cell(index+row_start ,col_start+11).font = Font(color="ffffff", size=2)
                    modi= 'False'
                    if row['MYMODIFIED']:
                        modi='True'
                    
                    worksheet.cell(index+row_start ,col_start+12 ).value =modi
                    worksheet.cell(index+row_start ,col_start+12).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+13 ).value =row['MYREADONLY']
                    worksheet.cell(index+row_start ,col_start+13).font = Font(color="ffffff", size=2)




                    

                    worksheet.cell(index+row_start ,col_start+6).value= row['MYSCORE']
                    worksheet.cell(index+row_start ,col_start+6).font=bold_font 
                    worksheet.cell(index+row_start ,col_start+6).alignment = center_aligned_text
                    worksheet.cell(index+row_start ,col_start+6).protection = Protection(locked=False)  
              
                worksheet.cell( len(courselist)+row_start  ,1).value='END'
                worksheet.cell(len(courselist)+row_start  ,1).font = Font(color="ffffff", size=2)
                worksheet.cell(len(courselist)+row_start  ,1).alignment = center_aligned_text
                workbook.save(response)
                return response
       else:
                print('Nothing was passed in session')
  else :
        print('This is a POST message')
        return render (request,'GradeManager/displayCourse_view.html',{'courselist':courselist})

@login_required
def  uploadScoresheet_xls(request):
    context={}
    if request.method == 'POST' :
        
        form = UploadedScoreForm(request.POST,request.FILES)
        if form.is_valid():
            
            excel_file = request.FILES['scoresheetfile']
            thename = (request.user.username).replace('@','').replace('/','')
            fs = FileSystemStorage()
            print(thename)
            filename = fs.save(thename+'-'+excel_file.name, excel_file)
            uploaded_file_url = fs.url(filename)



            try :
                lists,basicunits = generatescorelist(excel_file)
                print("The basics")
                print(basicunits)

                
                err,msg = validatelist(lists)
                if err == -1 :
                    messages.error(request, msg)
                    return render(request,'GradeManagerapp/uploadScoresheet_xls.html',context)
                #json_string = json.dumps([ob.__dict__ for ob in lists])
                json_string = [ob.__dict__ for ob in lists]
            
                api=settings.BASE_URL+'/api/Student/PythonUploadScore'
                print(api)
                data = MergebasicScorelist(basicunits.__dict__,json_string)
                mydata = json.dumps(data.__dict__)
            
            except  Exception as inst:
               print(excel_file.name+'  Has Error below ') 
               print(inst) 
               messages.error(request, "Problem With the Excel Score File Uploaded")
               return render(request,'GradeManagerapp/uploadScoresheet_xls.html',context)
                 

        try:
            headers = {'content-type': 'application/json'}
            
            r = requests.post(api,data=mydata,headers=headers)
            
            
            if r.status_code==200:
                messages.success(request, "  Successfully Uploaded")
            else:
                messages.error(request,str(r.status_code) +" Problem loading data")
        except  Exception as inst:
            print("See Error Details Below /n")
            print(inst)
            messages.error(request,inst)


    else :
        form = UploadedScoreForm()
    context={'form':form}
    return render(request,'GradeManagerapp/uploadScoresheet_xls.html',context)




@login_required
def downloadPdfReports(request):

    print('Beginning file downloadPdfReports with requests')
    api=settings.BASE_URL+'/api/Student/PythonPullForPdfReports'
    print(api)
    print(request.POST)
    deptcode= request.POST['deptcode']
    sessioncode= request.POST['sessioncode']
    semestercode= request.POST['semestercode']
    progtypecode= request.POST['progtypecode']
    progcode= request.POST['progcode']
    setcode= request.POST['setcode']
    reportname= request.POST['reportname']
    myLevelTodo= request.POST['myLevelTodo']
    #progcode= request.POST['progcode']
    # params = {'includescore':includescore,'longerreporttype':crsid,'orderbymatricno':orderbymatricno,'reportname':reportname,'step':step,'year':year,'month':month,'day':day}
    params = { 'myLevelTodo':myLevelTodo, 'reportname':reportname,'longerreporttype':'TRUE','mycampId':'IBA','myProgId':progcode,'myProgOptionId':deptcode,'myAsetId':setcode,'myAsessionId':sessioncode,'mySemesterId':semestercode,'progtypeId':progtypecode}
  
    filename=(reportname+myLevelTodo+progcode+deptcode+setcode+sessioncode+semestercode+progtypecode).replace("/", "_")
    print(filename)
    headers = {'content-type': 'application/json'}
    
    r = requests.post(api,json=params,headers=headers)
    response = HttpResponse(r.content,content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename='+filename+".pdf"
    return response

@login_required
def downloadScoreSheetPdf(request):

    print('Beginning file download with requests')
    api=settings.BASE_URL+'/api/Student/PythonPullForScoreSheetPdf'
    print(api)
    params = request.session['params']
    print(params)
    print("params")
    headers = {'content-type': 'application/json'}
    r = requests.post(api,json=params,headers=headers)
    response = HttpResponse(r.content,content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename='+params['longerreporttype']+".pdf"
    return response
